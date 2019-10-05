import datetime
import os
import sys

if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import time, sys
import openpyxl


class UpdateTime(QThread):
    time_show = pyqtSignal(str)

    def run(self):
        while 1:
            a = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.time_show.emit(str(a))
            time.sleep(1)


class UpdateSpend(QThread):
    time_spend = pyqtSignal(str)

    def run(self):
        spend = 0
        while 1:
            spend += 1
            h = int(spend / 60 / 60)
            m = int(spend / 60)
            s = spend % 60
            if h < 10:
                h = '0%d' % h
            if m < 10:
                m = '0%d' % m
            if s < 10:
                s = '0%d' % s
            str1 = '{}:{}:{}'.format(h, m, s)
            self.time_spend.emit(str(str1))
            time.sleep(1)


class LogWnd(QDialog):
    close_signal = pyqtSignal(str)

    _startPos = None
    _endPos = None
    _isTracking = False

    def __init__(self, start_time, end_time, spend_time, content='', ):
        QDialog.__init__(self)
        self.isSaved = False  # 判断在关闭窗口时是否需要提示保存对话框
        self.start_time = start_time  # 父窗口传递过来的开始时间
        self.end_time = end_time  # 父窗口传递过来的结束时间
        self.spend_time = spend_time.split(':', 1)[1]
        self.saved_text = ''
        self.selection = ''
        # load work list from excel
        self.process_list = []
        self.wait_list = []
        self.completed_list = []
        self.list_storys()
        # self.setFixedSize(440, 350)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)  # 无边框
        self.setStyleSheet(
            '''
            QTextEdit{
                border:none;
                color:black;
                background-color:white
                }
            QDialog{
                background-color:lightblue;
                border-top:1px solid white;
                border-bottom:1px solid white;
                border-left:1px solid white;
                border-right:1px solid white;
                border-top-left-radius:12px;
                border-bottom-left-radius:12px;
                border-top-right-radius:12px;
                border-bottom-right-radius:12px;
                }
            QComboBox{
                color:black;
                background:lightgreen;
                font-size:16px;
                text-align:center;
                border-top-left-radius:12px;
                border-bottom-left-radius:12px;
                border-top-right-radius:12px;
                border-bottom-right-radius:12px;
                }
            QLabel{
                border:none;
                color:blue;
                background:lightblue;
                font-size:16px;
                text-align:center;
                }
            QPushButton{
                border:none;
                color:white;
                background:gray;
                border:1px solid #F3F3F5;
                border-radius:10px;
                font-size:15px;
                height:40px;
                padding-left:10px;
                padding-right:10px;
                text-align:center;
                }
            QPushButton:hover{
                color:black;
                border:1px solid #F3F3F5;
                border-radius:10px;
                background:LightGray;
                }
                '''
        )

        # layout
        self.stoys = QComboBox(self)
        self.stoys.move(5, 10)
        self.stoys.resize(320, 25)
        for item in self.process_list:
            self.stoys.addItem(item)
        for item in self.wait_list:
            self.stoys.addItem(item)
        for item in self.completed_list:
            self.stoys.addItem(item)
        self.stoys.currentIndexChanged.connect(self.selection_changed)

        self.status = QComboBox(self)
        self.status.move(5, 40)
        self.status.resize(320, 25)
        self.status.addItems(['', 'Processing', 'Waitting', 'Completed'])

        self.log_field = QTextEdit(content, self)
        self.log_field.setGeometry(2, 70, 436, 307)
        self.log_field.setText(self.stoys.currentText())

        self.log_btn = QPushButton('Save', self)
        self.log_btn.setFixedSize(70, 25)
        self.log_btn.setShortcut('shift+alt+s')
        self.log_btn.move(335, 10)

        self.set_status_btn = QPushButton('Set', self)
        self.set_status_btn.setFixedSize(70, 25)
        self.set_status_btn.move(335, 40)
        self.set_status_btn.clicked.connect(self.set_status)

        self.close_btn = QPushButton('X', self)
        self.close_btn.setShortcut('shift+alt+c')
        self.close_btn.setFixedSize(30, 30)
        self.close_btn.move(408, 2)
        self.close_btn.setStyleSheet(
            """
            QPushButton{
                border:none;
                color:white;
                background:lightblue;
                font-size:15px;
                height:40px;
                padding-left:10px;
                padding-right:10px;
                text-align:center;
                }
            QPushButton:hover{
                color:white;
                border:1px solid #F3F3F5;
                border-radius:10px;
                background:red;
                }
            """
        )

        self.log_btn.clicked.connect(self.save)
        self.close_btn.clicked.connect(self.closeEvent)
        # self.closeEvent()

    def set_status(self):
        project = self.stoys.currentText()
        status = self.status.currentText()
        wb = openpyxl.load_workbook('worklist.xlsx')
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            if project == ws.cell(row, 1).value:
                ws.cell(row, 3).value = status
                wb.save('worklist.xlsx')
                break
            else:
                continue

    def list_storys(self):
        if not os.path.exists('worklist.xlsx'):
            QMessageBox.information(None, 'Warning',
                                    'No work list with name worklist.xlsx was found please double check and copy to {}'.format(
                                        os.getcwd()), QMessageBox.Yes)
            return
        else:
            wb = openpyxl.load_workbook('worklist.xlsx')
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                project_name = ws.cell(row, 1).value
                project_status = ws.cell(row, 3).value
                if project_status.upper() == 'PROCESSING':
                    self.process_list.append(project_name)
                elif project_status.upper() == 'WAITTING':
                    self.wait_list.append(project_name)
                elif project_status.upper() == 'COMPLETED':
                    self.completed_list.append(project_name)

    def selection_changed(self):
        self.selection = self.stoys.currentText()
        self.log_field.setText(self.selection)
        if self.selection in self.process_list:
            self.status.setCurrentText('Processing')
        elif self.selection in self.wait_list:
            self.status.setCurrentText('Waitting')
        elif self.selection in self.completed_list:
            self.status.setCurrentText('Completed')
        else:
            pass

    def closeEvent(self, QCloseEvent):
        self.close_signal.emit('closed')
        self.close_event()

    def close_event(self):
        if self.saved_text == self.log_field.toPlainText():
            self.close()
        else:
            isSaved = QMessageBox.information(None, 'Warning', 'Logged content is not saved, do you want saved?',
                                              QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if isSaved == 16384:
                self.save()
                self.close()
            elif isSaved == 65536:
                self.saved_text = self.log_field.toPlainText()
                self.close()

    def save(self):
        if self.isSaved:
            if self.saved_text == self.log_field.toPlainText():
                return
        self.saved_text = self.log_field.toPlainText()
        if os.path.exists('logwork.xlsx'):
            wb = openpyxl.load_workbook('logwork.xlsx')
            ws = wb.active
            start_row = ws.max_row + 1
            ws.cell(start_row, 1).value = self.start_time
            ws.cell(start_row, 2).value = self.end_time
            ws.cell(start_row, 3).value = self.spend_time
            ws.cell(start_row, 4).value = self.selection
            ws.cell(start_row, 5).value = self.saved_text
            wb.save('logwork.xlsx')
            self.isSaved = True
        else:
            workboot = openpyxl.Workbook()
            ws = workboot.active
            ws.cell(1, 1).value = 'Begin Time'
            ws.cell(1, 2).value = 'End Time'
            ws.cell(1, 3).value = 'Spend Time'
            ws.cell(1, 4).value = 'Type'
            ws.cell(1, 5).value = 'Work Log'
            ws.cell(2, 1).value = self.start_time
            ws.cell(2, 2).value = self.end_time
            ws.cell(2, 3).value = self.spend_time
            ws.cell(2, 4).value = self.selection
            ws.cell(2, 5).value = self.saved_text
            workboot.save('logwork.xlsx')
            self.isSaved = True

    def mouseMoveEvent(self, e: QMouseEvent):  # 重写移动事件
        try:
            self._endPos = e.pos() - self._startPos
            self.move(self.pos() + self._endPos)
        except:
            pass

    def mousePressEvent(self, e: QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._isTracking = True
            self._startPos = QPoint(e.x(), e.y())

    def mouseReleaseEvent(self, e: QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._isTracking = False
            self._startPos = None
            self._endPos = None


class Main(QDialog):
    _startPos = None
    _endPos = None
    _isTracking = False

    def __init__(self):
        QDialog.__init__(self)
        self.desktop = QApplication.desktop()
        self.screen_rect = self.desktop.screenGeometry()
        width, height = self.screen_rect.width(), self.screen_rect.height()

        self.setWindowTitle('字体设置')
        self.setFixedSize(QSize(305, 80))
        self.move(width - 350, 10)
        # border-image:url(img.jpg);
        self.setStyleSheet(
            '''
            QFrame{
                border:none;
                color:white;
                background:lightblue;
                border:1px solid #F3F3F5;
                border-radius:10px;
                text-align:center;
                }
            QDialog{
                background-color:lightgreen;
                border-top:1px solid white;
                border-bottom:1px solid white;
                border-left:1px solid white;
                border-right:1px solid white;
                border-top-left-radius:12px;
                border-bottom-left-radius:12px;
                border-top-right-radius:12px;
                border-bottom-right-radius:12px;
                }
            QLabel{
                border:none;
                color:black;
                background:lightgreen;
                border:1px solid #F3F3F5;
                border-radius:10px;
                font-size:12px;
                height:40px;
                padding-left:10px;
                padding-right:10px;
                text-align:center;
                }
            QPushButton{
                border:none;
                color:white;
                background:gray;
                border:1px solid #F3F3F5;
                border-radius:10px;
                font-size:15px;
                height:40px;
                padding-left:10px;
                padding-right:10px;
                text-align:center;
                }
            QPushButton:hover{
                color:black;
                border:1px solid #F3F3F5;
                border-radius:10px;
                background:LightGray;
                }
                '''
        )
        self.setAutoFillBackground(True)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog | Qt.WindowStaysOnTopHint)  # 无边框 置顶
        self.setAttribute(Qt.WA_TranslucentBackground)  # 透明度，去掉圆弧边角
        # layout
        self.label_frame = QFrame(self)
        self.label_frame.resize(305, 75)

        layout = QGridLayout(self.label_frame)
        self.time_label = QLabel('Current:', self.label_frame)
        self.time_label.resize(230, 20)
        self.time_spend = QLabel('Spend  :00:00:00', self.label_frame)
        self.time_spend.resize(230, 20)
        self.timer_btn = QPushButton('Start', self.label_frame)
        self.timer_btn.setShortcut('shift+alt+1')
        self.timer_btn.setFixedSize(70, 25)
        self.timer_close = QPushButton('Close', self.label_frame)
        self.timer_close.setShortcut('shift+alt+2')
        self.timer_close.setFixedSize(70, 25)
        self.timer_btn.clicked.connect(self.start_action)
        self.timer_close.clicked.connect(self.closeW)

        layout.addWidget(self.time_label, 0, 0)
        layout.addWidget(self.timer_btn, 0, 1)
        layout.addWidget(self.time_spend, 1, 0)
        layout.addWidget(self.timer_close, 1, 1)

        self.thread = UpdateTime()
        self.thread.time_show.connect(self.display_time)
        self.thread.start()

    def closeW(self):
        try:
            self.logwnd.close()
        except:
            pass
        self.close()

    def start_action(self):
        if not os.path.exists('worklist.xlsx'):
            QMessageBox.information(None, 'Warning', 'No work list with name worklist.xlsx was found please double'
                                                     ' check and copy to {}'.format(os.getcwd()), QMessageBox.Yes)
            return
        if self.timer_btn.text() == "Start":
            self.timer_btn.setText('Stop')
            self.timer_btn.setShortcut('shift+alt+1')
            self.begin_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            self.thread = UpdateSpend()
            self.thread.time_spend.connect(self.display_spend)
            self.thread.start()

        elif self.timer_btn.text() == "Stop":
            self.end_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.timer_btn.setText('Start')
            self.timer_btn.setShortcut('shift+alt+1')
            self.timer_btn.setEnabled(False)
            self.logwnd = LogWnd(start_time=self.begin_time, end_time=self.end_time,
                                 spend_time=self.time_spend.text())  # 加了self之后打开子窗口会停留，否则子窗口会一闪而过
            self.logwnd.close_signal.connect(self.update_close)
            self.logwnd.show()
            try:
                self.thread.disconnect()
            except:
                pass

    def update_close(self, msg):
        if msg == 'closed':
            self.timer_btn.setEnabled(True)
            self.time_spend.setText('Spend  :00:00:00')

    def display_time(self, msg):
        self.time_label.setText("Current:" + msg)

    def display_spend(self, msg):
        self.time_spend.setText('Spend  :' + msg)

    def mouseMoveEvent(self, e: QMouseEvent):  # 重写移动事件
        try:
            self._endPos = e.pos() - self._startPos
            self.move(self.pos() + self._endPos)
            e.accept()
        except:
            pass

    def mousePressEvent(self, e: QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._isTracking = True
            self._startPos = QPoint(e.x(), e.y())

    def mouseReleaseEvent(self, e: QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._isTracking = False
            self._startPos = None
            self._endPos = None


if __name__ == '__main__':
    app = QApplication(sys.argv)
    x = Main()
    x.show()
    sys.exit(app.exec_())
