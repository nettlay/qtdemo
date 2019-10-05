import datetime
import os
import sys
if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import time, sys
import openpyxl

class LogWnd(QDialog):
  close_signal = pyqtSignal(str)
  popupAboutToBeShown = pyqtSignal()

  _startPos = None
  _endPos = None
  _isTracking = False

  def __init__(self, start_time, end_time, spend_time, content='', ):
    QDialog.__init__(self)
    self.isSaved = False  # 判断在关闭窗口时是否需要提示保存对话框
    self.start_time = start_time  # 父窗口传递过来的开始时间
    self.end_time = end_time  # 父窗口传递过来的结束时间
    self.spend_time = spend_time.split(':', 1)[1]
    self.saved_text = ''
    # self.setFixedSize(440, 350)
    self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)  # 无边框
    self.setStyleSheet(
      '''
      QTextEdit{
          border:none;
          color:black;
          background-color:white
          }
      QDialog{
          background-color:lightblue;
          border-top:1px solid white;
          border-bottom:1px solid white;
          border-left:1px solid white;
          border-right:1px solid white;
          border-top-left-radius:12px;
          border-bottom-left-radius:12px;
          border-top-right-radius:12px;
          border-bottom-right-radius:12px;
          }
      QLabel{
          border:none;
          color:blue;
          background:lightblue;
          font-size:16px;
          text-align:center;
          }
      QPushButton{
          border:none;
          color:white;
          background:gray;
          border:1px solid #F3F3F5;
          border-radius:10px;
          font-size:20px;
          height:40px;
          padding-left:10px;
          padding-right:10px;
          text-align:center;
          }
      QPushButton:hover{
          color:black;
          border:1px solid #F3F3F5;
          border-radius:10px;
          background:LightGray;
          }
          '''
    )
    self.log_field = QTextEdit(content, self)
    self.log_field.setGeometry(2, 40, 436, 307)

    # layout
    self.stoys = QComboBox(self)
    self.stoys.move(5, 10)
    self.stoys.resize(300, 25)
    self.stoys.addItem('hhhh')
    self.stoys.addItems(['test1', 'test2'])
    self.stoys.currentIndexChanged.connect(self.selection_changed)

    self.log_btn = QPushButton('Save', self)
    self.log_btn.setFixedSize(70, 30)
    self.log_btn.setShortcut('shift+alt+s')
    self.log_btn.move(330, 5)
    self.close_btn = QPushButton('X', self)
    self.close_btn.setShortcut('shift+alt+c')
    self.close_btn.setFixedSize(30, 30)
    self.close_btn.move(405, 5)

    self.log_btn.clicked.connect(self.save)
    self.close_btn.clicked.connect(self.closeEvent)

    # self.closeEvent()

  def selection_changed(self, selection):
    print(self.stoys.currentText())
    print(selection)

  def closeEvent(self, QCloseEvent):
    self.close_signal.emit('closed')
    self.close_event()

  def close_event(self):
    if self.saved_text == self.log_field.toPlainText():
      self.close()
    else:
      isSaved = QMessageBox.information(None, 'Warning', 'Logged content is not saved, do you want saved?',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
      if isSaved == 16384:
        self.save()
        self.close()
      elif isSaved == 65536:
        self.saved_text = self.log_field.toPlainText()
        self.close()

  def save(self):
    if self.isSaved:
      if self.saved_text == self.log_field.toPlainText():
        return
    self.saved_text = self.log_field.toPlainText()
    if os.path.exists('logwork.xlsx'):
      wb = openpyxl.load_workbook('logwork.xlsx')
      ws = wb.active
      start_row = ws.max_row + 1
      ws.cell(start_row, 1).value = self.start_time
      ws.cell(start_row, 2).value = self.end_time
      ws.cell(start_row, 3).value = self.spend_time
      ws.cell(start_row, 4).value = 'None'
      ws.cell(start_row, 5).value = self.saved_text
      wb.save('logwork.xlsx')
      self.isSaved = True
    else:
      workboot = openpyxl.Workbook()
      ws = workboot.active
      ws.cell(1, 1).value = 'Begin Time'
      ws.cell(1, 2).value = 'End Time'
      ws.cell(1, 3).value = 'Spend Time'
      ws.cell(1, 4).value = 'Type'
      ws.cell(1, 5).value = 'Work Log'
      ws.cell(2, 1).value = self.start_time
      ws.cell(2, 2).value = self.end_time
      ws.cell(2, 3).value = self.spend_time
      ws.cell(2, 4).value = 'None'
      ws.cell(2, 5).value = self.saved_text
      workboot.save('logwork.xlsx')
      self.isSaved = True

  def mouseMoveEvent(self, e: QMouseEvent):  # 重写移动事件
    try:
      self._endPos = e.pos() - self._startPos
      self.move(self.pos() + self._endPos)
      e.accept()
    except:
      pass

  def mousePressEvent(self, e: QMouseEvent):
    if e.button() == Qt.LeftButton:
      self._isTracking = True
      self._startPos = QPoint(e.x(), e.y())

  def mouseReleaseEvent(self, e: QMouseEvent):
    if e.button() == Qt.LeftButton:
      self._isTracking = False
      self._startPos = None
      self._endPos = None


if __name__ == '__main__':
    app = QApplication(sys.argv)
    demo = LogWnd('af:df', 'a:aa', '2:3s')
    demo.show()
    sys.exit(app.exec_())